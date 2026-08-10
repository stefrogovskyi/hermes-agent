# -*- coding: utf-8 -*-
"""
build_navo_excel_report.py — Генератор Excel-отчёта по продажам Navo24 и 2-месячному плану.
Формирует 4 вкладки:
  1. Summary Dashboard (Сводные графики, метрики и общая выручка)
  2. Manager Portfolios (Детальная разбивка по всем 8 менеджерам по именам)
  3. Renewal Timeline & Risk Audit (Хронологический календарь продлений Q3/Q4 2026 - 2027)
  4. 2-Month Action Plan (Пошаговый план работ на Август-Сентябрь 2026)
"""

import os, sys, openpyxl
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPORTS_DIR = r"C:\Users\Stefan\AppData\Local\hermes\reports"
os.makedirs(REPORTS_DIR, exist_ok=True)

excel_path = os.path.join(REPORTS_DIR, "navo_sales_team_audit_and_2month_plan_20260802.xlsx")

wb = openpyxl.Workbook()

# Styling
title_font = Font(name="Segoe UI", size=16, bold=True, color="0284C7")
subtitle_font = Font(name="Segoe UI", size=11, italic=True, color="64748B")
section_font = Font(name="Segoe UI", size=12, bold=True, color="1E293B")
header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
bold_font = Font(name="Segoe UI", size=11, bold=True, color="000000")
regular_font = Font(name="Segoe UI", size=11, color="000000")

blue_header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
dark_header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
yellow_fill = PatternFill(start_color="FEF08A", end_color="FEF08A", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

# ---------------------------------------------------------------------------
# Sheet 1: Summary Dashboard
# ---------------------------------------------------------------------------
ws1 = wb.active
ws1.title = "Summary Dashboard"

ws1.cell(row=1, column=1, value="Navo24 Sales Team Revenue Audit & 2-Month Execution Plan").font = title_font
ws1.cell(row=2, column=1, value=f"Director: Stefan Rogovskiy | Date: {datetime.now().strftime('%d.%m.%Y')} | Target: $100k+ MRR Guaranteed").font = subtitle_font

# Key KPI Cards
kpis = [
    ("Ежемесячный Cash-Flow (PAYG + Monthly)", "$147,122 / мес", "Текущий фактический ежемесячный поток"),
    ("Портфель Годовых Контрактов (ARR)", "$2,384,223 / год", "$198,685 / мес в амортизации"),
    ("Совокупный Эквивалент Выручки", "$345,807 / мес", "Общий объём бизнеса в месяц"),
    ("Количество Сейлз-Менеджеров", "8 Человек", "Катя К., Катя К., Саша, Андрей, Олег, Катя К., Лера, Лиля")
]

for idx, (kpi_t, kpi_v, kpi_d) in enumerate(kpis, 1):
    c_col = (idx - 1) * 2 + 1
    ws1.cell(row=4, column=c_col, value=kpi_t).font = Font(bold=True, color="475569")
    v_cell = ws1.cell(row=5, column=c_col, value=kpi_v)
    v_cell.font = Font(size=14, bold=True, color="0284C7")
    ws1.cell(row=6, column=c_col, value=kpi_d).font = Font(size=9, italic=True, color="64748B")

ws1.cell(row=8, column=1, value="СВОДНАЯ ТАБЛИЦА ВЫРУЧКИ ПО МЕНЕДЖЕРАМ").font = section_font

headers1 = ["Менеджер", "Помесячный Поток (PAYG/Mo)", "Годовой Портфель (ARR)", "Амортизация в Месяц", "Ключевые Клиенты / Контракты", "Главный Риск / Срок"]
for c_i, h in enumerate(headers1, 1):
    c = ws1.cell(row=9, column=c_i, value=h)
    c.fill = blue_header_fill
    c.font = header_font

summary_data = [
    ("Катя Комарова", 50000, 600000, 50000, "Топ-3 PAYG ($50k/mo)", "Договор до Sept 2026 и 2027"),
    ("Саша Грабарчук", 58844, 700000, 58333, "$38.9k/mo + $15.1k/mo + $4.6k/mo", "Октябрь 2026 ($38.9k/mo)"),
    ("Лера Гулий", 13500, 300000, 25000, "Помесячные $12-15k/mo", "03.10.2026 ($300k ARR)"),
    ("Лиля Ховрак", 10000, 500000, 41667, "Помесячные $10k/mo", "Крупный годовой портфель $500k"),
    ("Андрей Городинский", 11000, 92000, 7667, "Nauta ($6.5k), First ($2.2k), Narwal ($2.2k)", "Herpot ($15k/qtr Sept 2026)"),
    ("Катя Кернеш", 1300, 100790, 8399, "1 клиент ($1.3k/mo)", "14.11.2026 ($57.6k ARR)"),
    ("Олег Червинский", 1478, 39433, 3286, "Tradewind ($4.4k/3mo test), Perdue, Nauffar", "Конец августа 2026 (Tradewind test)"),
    ("Катя Капустян", 1000, 52000, 4333, "Малюки (<$1k/mo)", "07.04.2027 ($52k от Малахова)")
]

for r_i, r_vals in enumerate(summary_data, 10):
    ws1.cell(row=r_i, column=1, value=r_vals[0]).font = bold_font
    ws1.cell(row=r_i, column=2, value=r_vals[1]).number_format = "$#,##0"
    ws1.cell(row=r_i, column=3, value=r_vals[2]).number_format = "$#,##0"
    ws1.cell(row=r_i, column=4, value=r_vals[3]).number_format = "$#,##0"
    ws1.cell(row=r_i, column=5, value=r_vals[4]).font = regular_font
    ws1.cell(row=r_i, column=6, value=r_vals[5]).font = regular_font
    for c_i in range(1, 7):
        ws1.cell(row=r_i, column=c_i).border = thin_border

tot_row = 10 + len(summary_data)
ws1.cell(row=tot_row, column=1, value="ИТОГО КОМАНДА:").font = bold_font
ws1.cell(row=tot_row, column=2, value="=SUM(B10:B17)").number_format = "$#,##0"
ws1.cell(row=tot_row, column=3, value="=SUM(C10:C17)").number_format = "$#,##0"
ws1.cell(row=tot_row, column=4, value="=SUM(D10:D17)").number_format = "$#,##0"
for c_i in range(1, 7):
    c = ws1.cell(row=tot_row, column=c_i)
    c.font = bold_font
    c.border = thin_border
    c.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

# ---------------------------------------------------------------------------
# Sheet 2: Manager Portfolios
# ---------------------------------------------------------------------------
ws2 = wb.create_sheet("Manager Portfolios")
ws2.cell(row=1, column=1, value="Детальная разбивка портфелей по всем 8 менеджерам").font = title_font

headers2 = ["Менеджер", "Тип Оплаты", "Клиент / Продукт", "Сумма в Месяц", "Сумма в Год", "Срок / Дата Продления", "Примечания & Стратегия"]
for c_i, h in enumerate(headers2, 1):
    c = ws2.cell(row=3, column=c_i, value=h)
    c.fill = blue_header_fill
    c.font = header_font

details = [
    ("Катя Комарова", "PAYG / Помесячно", "Топ-3 Клиента (суммарно)", 50000, 600000, "Sept 2026 / 2027", "Обязательства по договору. Фиксация минимального тира."),
    ("Катя Капустян", "Помесячно", "Малюки (<$1k/mo)", 1000, 12000, "Текущие", "Апсейлл до пакетной подписки $1.5k/mo."),
    ("Катя Капустян", "Годовой", "Клиент 1 (от Малахова)", 2166, 26000, "07.04.2027", "Стабильный годовой контракт."),
    ("Катя Капустян", "Годовой", "Клиент 2 (от Малахова)", 2166, 26000, "19.06.2027", "Стабильный годовой контракт."),
    ("Саша Грабарчук", "Помесячно", "Клиент 1 (Top PAYG)", 38992, 467904, "October 2026", "КРИТИЧЕСКИЙ РИСК: Проведение QBR в сентябре для продления."),
    ("Саша Грабарчук", "Помесячно", "Клиент 2 (2 сервиса)", 15169, 182028, "Nov - Dec 2026", "Высокая вовлеченность (2 сервиса)."),
    ("Саша Грабарчук", "Помесячно", "Клиент 3", 4683, 56196, "August 2027", "Долгосрочный контракт."),
    ("Андрей Городинский", "PAYG", "Nauta", 6500, 78000, "PAYG (5-8k/mo)", "Зафиксировать фикс-тир $6.5k/mo."),
    ("Андрей Городинский", "PAYG", "First", 2250, 27000, "PAYG (2-2.5k/mo)", "Зафиксировать фикс-тир $2.25k/mo."),
    ("Андрей Городинский", "PAYG", "Narwal", 2250, 27000, "PAYG (2-2.5k/mo)", "Зафиксировать фикс-тир $2.25k/mo."),
    ("Андрей Городинский", "SaaS Годовой", "Tadanow", 1250, 15000, "Март-Апрель 2027", "Стабильный SaaS."),
    ("Андрей Городинский", "SaaS Квартал", "Herpot", 5000, 60000, "Сентябрь 2026 ($15k/qtr)", "Продление Q4 в сентябре или перевод на $50k/yr."),
    ("Андрей Городинский", "SaaS Годовой", "Indofood", 1416, 17000, "Январь 2027", "Крупный азиатский фуд-гигант."),
    ("Олег Червинский", "Тест 3 мес", "Tradewind International", 1478, 17744, "Конец Августа 2026 ($4.4k/3mo)", "ГЛАВНАЯ ТОЧКА РОСТА АВГУСТА: Конверсия в $20k/yr."),
    ("Олег Червинский", "Годовой API", "Perdue AgriBusiness", 455, 5460, "01.07.2027", "Агро-сектор."),
    ("Олег Червинский", "Годовой API", "Nauffar Germany", 240, 2880, "Конец Июля 2027", "Европейский логист."),
    ("Олег Червинский", "Годовой Air", "Escavox South Africa", 916, 11000, "Годовой", "Авиа-трекинг (от Жени)."),
    ("Олег Червинский", "Годовой", "Regenerate Trade & Log", 195, 2349, "Январь 2027", "API + Web-интеграция."),
    ("Катя Кернеш", "Помесячно", "Клиент 1", 1300, 15600, "Текущий", "Апсейлл до $2.5k/mo."),
    ("Катя Кернеш", "Годовой", "VIP Клиент 1", 4800, 57600, "Nov 14, 2026", "Крупный годовой контракт."),
    ("Катя Кернеш", "Годовой", "VIP Клиент 2", 1999, 23990, "Jul 23, 2027", "Годовой контракт."),
    ("Катя Кернеш", "Годовой", "VIP Клиент 3", 1600, 19200, "Jun 25, 2027", "Годовой контракт."),
    ("Лера Гулий", "Помесячно", "Топ-3 Помесячные", 13500, 162000, "Текущие ($12-15k/mo)", "Защита объема."),
    ("Лера Гулий", "Годовой", "Топ-3 Годовые", 25000, 300000, "03.10.2026 / 13.12.2026 / 01.01.2027", "КРИТИЧЕСКИЙ РИСК: $300k ARR продление в октябре."),
    ("Лиля Ховрак", "Помесячно", "Топ-3 Помесячные", 10000, 120000, "Текущие (~$10k/mo)", "Защита объема."),
    ("Лиля Ховрак", "Годовой", "Топ-3 Годовые", 41667, 500000, "Годовые (~$500k ARR)", "Крупнейший портфель.")
]

for r_i, r_vals in enumerate(details, 4):
    ws2.cell(row=r_i, column=1, value=r_vals[0]).font = bold_font
    ws2.cell(row=r_i, column=2, value=r_vals[1]).font = regular_font
    ws2.cell(row=r_i, column=3, value=r_vals[2]).font = bold_font
    ws2.cell(row=r_i, column=4, value=r_vals[3]).number_format = "$#,##0"
    ws2.cell(row=r_i, column=5, value=r_vals[4]).number_format = "$#,##0"
    ws2.cell(row=r_i, column=6, value=r_vals[5]).font = regular_font
    ws2.cell(row=r_i, column=7, value=r_vals[6]).font = regular_font
    for c_i in range(1, 8):
        ws2.cell(row=r_i, column=c_i).border = thin_border

# ---------------------------------------------------------------------------
# Sheet 3: Renewal Timeline & Risk Audit
# ---------------------------------------------------------------------------
ws3 = wb.create_sheet("Renewal Timeline")
ws3.cell(row=1, column=1, value="Хронологический Календарь Продлений & Аудит Рисков (Q3/Q4 2026 - 2027)").font = title_font

headers3 = ["Дата Продления", "Менеджер", "Клиент / Продукт", "Объем Суммы", "Уровень Риска", "Стратегическое Действие"]
for c_i, h in enumerate(headers3, 1):
    c = ws3.cell(row=3, column=c_i, value=h)
    c.fill = blue_header_fill
    c.font = header_font

timeline_data = [
    ("Конец Августа 2026", "Олег Червинский", "Tradewind International", "$4,436 / 3 мес", "ВЫСОКИЙ (Завершение теста)", "Предоставить отчет точности SeaRates vs TrackingMCP; конвертировать в $20k/yr годовой."),
    ("Сентябрь 2026", "Катя Комарова", "Клиент 1 (PAYG)", "$50,000 / мес", "СРЕДНИЙ (Обязательство)", "Перевод на годовой контракт со скидкой 10%."),
    ("Сентябрь 2026", "Андрей Городинский", "Herpot", "$15,000 / квартал", "СРЕДНИЙ (Квартальное продление)", "Продлить на Q4 или перевести на $50k/yr upfront."),
    ("03.10.2026", "Лера Гулий", "Крупный Годовой Клиент", "$300,000 / год", "КРИТИЧЕСКИЙ (Годовой пролонгат)", "Провести QBR в сентябре, показать ROI за год и досрочно переподписать."),
    ("Октябрь 2026", "Саша Грабарчук", "Крупнейший PAYG Клиент", "$38,992 / мес", "КРИТИЧЕСКИЙ (Крупнейший доход)", "Провести технический аудит в сентябре, зафиксировать годовой $450k+ контракт."),
    ("14.11.2026", "Катя Кернеш", "VIP Клиент 1", "$57,600 / год", "СРЕДНИЙ", "Предложить пакетное добавление LoadingMCP и SchedulesMCP."),
    ("13.12.2026", "Лера Гулий", "Годовой Клиент 2", "$300k портфель", "СРЕДНИЙ", "Подготовка QBR за 30 дней."),
    ("01.01.2027", "Лера Гулий", "Годовой Клиент 3", "$300k портфель", "СРЕДНИЙ", "Подготовка QBR за 30 дней."),
    ("Январь 2027", "Андрей Городинский", "Indofood", "$17,000 / год", "НИЗКИЙ", "Автоматическое продление."),
    ("Январь 2027", "Олег Червинский", "Regenerate Trade", "$2,349 / год", "НИЗКИЙ", "Автоматическое продление."),
    ("Март-Апрель 2027", "Андрей Городинский", "Tadanow", "$15,000 / год", "НИЗКИЙ", "Автоматическое продление."),
    ("07.04.2027", "Катя Капустян", "Клиент 1 (от Малахова)", "$26,000 / год", "НИЗКИЙ", "Автоматическое продление."),
    ("19.06.2027", "Катя Капустян", "Клиент 2 (от Малахова)", "$26,000 / год", "НИЗКИЙ", "Автоматическое продление."),
    ("25.06.2027", "Катя Кернеш", "VIP Клиент 3", "$19,200 / год", "НИЗКИЙ", "Автоматическое продление."),
    ("01.07.2027", "Олег Червинский", "Perdue AgriBusiness", "$5,460 / год", "НИЗКИЙ", "Автоматическое продление."),
    ("23.07.2027", "Катя Кернеш", "VIP Клиент 2", "$23,990 / год", "НИЗКИЙ", "Автоматическое продление."),
    ("Конец Июля 2027", "Олег Червинский", "Nauffar Germany", "$2,880 / год", "НИЗКИЙ", "Автоматическое продление.")
]

for r_i, r_vals in enumerate(timeline_data, 4):
    ws3.cell(row=r_i, column=1, value=r_vals[0]).font = bold_font
    ws3.cell(row=r_i, column=2, value=r_vals[1]).font = regular_font
    ws3.cell(row=r_i, column=3, value=r_vals[2]).font = bold_font
    ws3.cell(row=r_i, column=4, value=r_vals[3]).font = regular_font
    
    r_cell = ws3.cell(row=r_i, column=5, value=r_vals[4])
    r_cell.font = Font(bold=True, color="DC2626" if "КРИТИЧЕСКИЙ" in r_vals[4] or "ВЫСОКИЙ" in r_vals[4] else "059669")
    
    ws3.cell(row=r_i, column=6, value=r_vals[5]).font = regular_font
    for c_i in range(1, 7):
        ws3.cell(row=r_i, column=c_i).border = thin_border

# ---------------------------------------------------------------------------
# Sheet 4: 2-Month Action Plan (Aug-Sep)
# ---------------------------------------------------------------------------
ws4 = wb.create_sheet("2-Month Action Plan")
ws4.cell(row=1, column=1, value="Пошаговый Операционный План Выхода на Гарантированные $100k+ MRR (Август - Сентябрь 2026)").font = title_font

headers4 = ["Месяц", "Ответственный", "Ключевая Задача / Инициатива", "Целевой KPI", "Срок Выполнения", "Роль AI Субагентов (Hermes / Richard / Alistair)"]
for c_i, h in enumerate(headers4, 1):
    c = ws4.cell(row=3, column=c_i, value=h)
    c.fill = blue_header_fill
    c.font = header_font

plan_data = [
    ("Август 2026", "Олег Червинский", "Конверсия Tradewind ($4.4k/3mo test) в годовой контракт", "Контракт $20,000 / год", "25.08.2026", "Ричард формирует DCSA-сравнительный отчёт точности с SeaRates."),
    ("Август 2026", "Андрей Городинский", "Перевод Nauta ($6.5k), First ($2.2k), Narwal ($2.2k) на фикс-тир", "Фиксация $10,000/мес MRR", "20.08.2026", "Алистер настраивает автоматический контроль пролива объема в Navo Tasktracker."),
    ("Август 2026", "Все 8 Менеджеров", "Пакетный Кросс-сейлл (SchedulesMCP + LoadingMCP) по Топ-15 клиентам", "+$15,000/мес нового MRR", "31.08.2026", "AI Скаут находит клиентов без 3D-укладки и отправляет интерактивное демо."),
    ("Август 2026", "Катя Комарова", "Защита $50k/mo PAYG портфеля и оформление минимального гаранта", "Сохранение $50,000/мес", "31.08.2026", "Hermes готовит годовое соглашение со скидкой 10%."),
    ("Сентябрь 2026", "Андрей Городинский", "Продление Herpot ($15,000/квартал) на Q4 2026", "Подписание $15,000 или $50k/yr", "20.09.2026", "Ричард готовит квартальный отчёт использования API и экономии."),
    ("Сентябрь 2026", "Лера Гулий", "Проведение QBR и досрочное продление годового клиента на $300k", "Переподписание $300,000 ARR", "25.09.2026", "Hermes готовит презентацию ROI за 12 месяцев использование TrackingMCP."),
    ("Сентябрь 2026", "Саша Грабарчук", "Проведение QBR по клиенту $38,992/мес и фиксация годового контракта", "Фиксация $450k+ ARR", "25.09.2026", "Алистер и Каллум готовят инспекцию стабильности API за год."),
    ("Сентябрь 2026", "Все 8 Менеджеров", "Подключение 10 Сейлзов к авто-воронке Hermes (2 сделки на сейлза)", "+$100,000 нового ARR", "30.09.2026", "Лиз Харпер ведет учет бонусов (15% от продаж), Hermes Scout генерирует лиды.")
]

for r_i, r_vals in enumerate(plan_data, 4):
    ws4.cell(row=r_i, column=1, value=r_vals[0]).font = bold_font
    ws4.cell(row=r_i, column=2, value=r_vals[1]).font = bold_font
    ws4.cell(row=r_i, column=3, value=r_vals[2]).font = regular_font
    ws4.cell(row=r_i, column=4, value=r_vals[3]).font = bold_font
    ws4.cell(row=r_i, column=5, value=r_vals[4]).font = regular_font
    ws4.cell(row=r_i, column=6, value=r_vals[5]).font = regular_font
    for c_i in range(1, 7):
        ws4.cell(row=r_i, column=c_i).border = thin_border

# Auto-adjust column widths for all sheets
for sheet in wb.worksheets:
    for col in sheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 65)

wb.save(excel_path)
print(f"Successfully generated Excel report at: {excel_path}")
